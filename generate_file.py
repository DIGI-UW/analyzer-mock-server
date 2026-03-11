#!/usr/bin/env python3
"""
Generate CSV/TXT/XLS files from FILE protocol templates.

This script generates file-based analyzer results from analyzer templates
that specify "protocol.type": "FILE". It supports the astm-mock-server template system
for testing file-based analyzer plugins.

Usage:
    python3 generate_file.py --template hain_fluorocycler --output /tmp/test.csv --count 5
    python3 generate_file.py -t quantstudio7 -o /tmp/QS7.xls -c 5 --format xls
"""

from template_loader import TemplateLoader
import argparse
import random
import sys
from datetime import datetime


def generate_csv(template, output_path, count=5):
    """Generate CSV file from FILE protocol template.
    
    Args:
        template: Loaded template dictionary with fileFormat and columns
        output_path: Output file path
        count: Number of data rows to generate
    """
    # Validate protocol type
    protocol = template.get('protocol', {})
    if protocol.get('type') != 'FILE':
        print(f"Error: Template protocol type is '{protocol.get('type')}', expected 'FILE'", 
              file=sys.stderr)
        return False
    
    file_format = template.get('fileFormat', {})
    delimiter = file_format.get('delimiter', ';')
    # Sort columns by index to ensure correct output order
    columns = sorted(template.get('columns', []), key=lambda c: c.get('index', 0))
    fields = template.get('fields', [])
    test_samples = template.get('testSamples', [])

    layout = template.get('layout', 'standard')
    if layout == 'plate_grid':
        if not test_samples:
            print("Error: Plate grid template requires testSamples", file=sys.stderr)
            return False
    elif not columns:
        print("Error: Template has no columns defined", file=sys.stderr)
        return False
    
    lines = []

    # Metadata block (Tecan F50, Multiskan FC, etc.)
    metadata = template.get('metadataBlock', [])
    for item in metadata:
        lines.append(f"{item.get('key', '')}{delimiter}{item.get('value', '')}")
    if metadata:
        lines.append('')

    layout = template.get('layout', 'standard')
    if layout == 'plate_grid':
        return _generate_plate_grid(template, output_path, count, delimiter, columns,
                                    test_samples, file_format)

    # Write header if specified
    if file_format.get('hasHeader', True):
        header = delimiter.join([c['name'] for c in columns])
        lines.append(header)

    # Generate data rows - use testSamples if available, otherwise auto-generate
    if test_samples and len(test_samples) >= count:
        # Use deterministic test samples from template
        for i in range(count):
            sample = test_samples[i]
            values = []
            for col in columns:
                col_name = col['name'].lower()

                # Map column names to test sample fields
                if col_name == 'position':
                    values.append(sample.get('position', f"{chr(65 + (i // 12))}{(i % 12) + 1:02d}"))
                elif 'sample' in col_name and 'id' in col_name:
                    values.append(sample.get('sampleId', f"SIM-{datetime.now().strftime('%Y%m%d')}-{i+1:04d}"))
                elif col_name == 'result':
                    values.append(sample.get('result', ''))
                elif col_name == 'interpretation':
                    values.append(sample.get('interpretation', ''))
                else:
                    # Try to find value in sample by column name
                    values.append(sample.get(col['name'], ''))

            lines.append(delimiter.join(values))
    else:
        # Auto-generate samples
        for i in range(count):
            sample_id = f"SIM-{datetime.now().strftime('%Y%m%d')}-{i+1:04d}"
            # Generate position: A01, A02, ..., A12, B01, B02, etc.
            position = f"{chr(65 + (i // 12))}{(i % 12) + 1:02d}"

            # Generate field values for each column
            values = []
            for col in columns:
                col_name = col['name']

                # Handle special column names
                if col_name.lower() == 'position':
                    values.append(position)
                elif 'sample' in col_name.lower() and 'id' in col_name.lower():
                    values.append(sample_id)
                else:
                    # Find matching field and pick random value from possibleValues
                    value_found = False
                    for field in fields:
                        # Match by code or name (case-insensitive partial match)
                        if (field['code'].lower() in col_name.lower() or
                            col_name.lower() in field.get('name', '').lower()):
                            possible = field.get('possibleValues', ['Unknown'])
                            values.append(random.choice(possible))
                            value_found = True
                            break

                    if not value_found:
                        # Default to empty string for unmatched columns
                        values.append('')

            lines.append(delimiter.join(values))
    
    # Write to file with specified encoding
    encoding = file_format.get('encoding', 'UTF-8')
    try:
        with open(output_path, 'w', encoding=encoding) as f:
            f.write('\n'.join(lines))
            # Add trailing newline if the last line doesn't have one
            if not lines[-1].endswith('\n'):
                f.write('\n')
        
        print(f"✓ Generated {count} samples to {output_path}")
        print(f"  Template: {template.get('analyzer', {}).get('name', 'Unknown')}")
        print(f"  Delimiter: '{delimiter}'")
        print(f"  Columns: {len(columns)}")
        return True
        
    except IOError as e:
        print(f"Error writing file: {e}", file=sys.stderr)
        return False


def generate_xls(template, output_path, count=5):
    """Generate XLS (BIFF8) file from QuantStudio-style template.

    Matches quantstudio-field-mapping-spec-v131: Results sheet, metadata block,
    header row with "Well" as first column, 31-column layout. Two rows per well
    (VIH-1 target + IC internal control).
    """
    try:
        import xlwt
    except ImportError:
        print("Error: xlwt required for XLS generation. Install with: pip install xlwt",
              file=sys.stderr)
        return False

    protocol = template.get('protocol', {})
    if protocol.get('type') != 'FILE':
        print(f"Error: Template protocol type is '{protocol.get('type')}', expected 'FILE'",
              file=sys.stderr)
        return False

    excel_format = template.get('excelFormat', {})
    columns = sorted(template.get('columns', []), key=lambda c: c.get('index', 0))
    test_samples = template.get('testSamples', [])

    if not columns:
        print("Error: Template has no columns defined", file=sys.stderr)
        return False

    wb = xlwt.Workbook()
    sheet_name = excel_format.get('sheetName', 'Results')
    ws = wb.add_sheet(sheet_name)

    row_idx = 0

    # Metadata block (key in col 0, value in col 1)
    metadata = excel_format.get('metadataBlock', [])
    for item in metadata:
        ws.write(row_idx, 0, item.get('key', ''))
        ws.write(row_idx, 1, item.get('value', ''))
        row_idx += 1

    # Pad to metadataRowCount if needed
    target_meta_rows = excel_format.get('metadataRowCount', len(metadata))
    while row_idx < target_meta_rows:
        ws.write(row_idx, 0, '')
        ws.write(row_idx, 1, '')
        row_idx += 1

    # Blank row before header
    row_idx += 1

    # Header row (first cell = "Well")
    for col_idx, col in enumerate(columns):
        ws.write(row_idx, col_idx, col['name'])
    row_idx += 1

    # Data rows: two per well (VIH-1 + IC) for each sample
    samples = test_samples[:count] if test_samples else []
    if not samples:
        for i in range(count):
            well_num = i + 1
            pos = f"{chr(65 + (i // 12))}{(i % 12) + 1:02d}"
            samples.append({
                'sampleId': f"LM{datetime.now().strftime('%Y%m%d')}{i+1:03d}",
                'position': pos,
                'task': 'UNKNOWN',
                'ct': 28.5 if i % 2 == 0 else 'Undetermined',
                'quantityMean': 1500 if i % 2 == 0 else ''
            })

    for sample in samples:
        pos = sample.get('position', 'A01')
        well_num = _well_position_to_index(pos)
        sample_id = sample.get('sampleId', '')
        task = sample.get('task', 'UNKNOWN')
        ct_val = sample.get('ct', '')
        qty_mean = sample.get('quantityMean', '')
        omit = sample.get('omit', 0)

        # VIH-1 row (target)
        _write_quantstudio_row(ws, row_idx, columns, well_num, pos, omit, sample_id,
                               'VIH-1', task, 'FAM', 'NFQ-MGB', ct_val, qty_mean)
        row_idx += 1

        # IC row (internal control)
        ic_ct = 25.0 if str(ct_val) == 'Undetermined' or ct_val == '' else 24.0
        _write_quantstudio_row(ws, row_idx, columns, well_num, pos, omit, sample_id,
                               'IC', task, 'CY5', 'None', ic_ct, '')
        row_idx += 1

    try:
        wb.save(output_path)
        print(f"✓ Generated {len(samples)} samples ({len(samples) * 2} data rows) to {output_path}")
        print(f"  Template: {template.get('analyzer', {}).get('name', 'Unknown')}")
        print(f"  Format: XLS (BIFF8), sheet: {sheet_name}")
        return True
    except IOError as e:
        print(f"Error writing file: {e}", file=sys.stderr)
        return False


def _generate_plate_grid(template, output_path, count, delimiter, columns,
                         test_samples, file_format):
    """Generate ELISA plate-grid format (8 rows x 12 cols) after metadata."""
    lines = []
    metadata = template.get('metadataBlock', [])
    for item in metadata:
        lines.append(f"{item.get('key', '')}{delimiter}{item.get('value', '')}")
    if metadata:
        lines.append('')

    # Grid header: <>	1	2	3	...	12
    header_cells = ['<>'] + [str(i) for i in range(1, 13)]
    lines.append(delimiter.join(header_cells))

    # Build 8x12 grid from testSamples or fill with placeholder values
    grid = [[0.0] * 12 for _ in range(8)]
    for i, sample in enumerate(test_samples[:min(96, count * 2)]):
        pos = sample.get('WellPosition', sample.get('position', ''))
        if not pos or len(pos) < 2:
            continue
        row = ord(pos[0].upper()) - ord('A')
        try:
            col = int(pos[1:]) - 1
        except ValueError:
            col = 0
        if 0 <= row < 8 and 0 <= col < 12:
            val = sample.get('OD_450', sample.get('result', 0.05))
            grid[row][col] = float(val) if val else 0.05

    for r, row_label in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']):
        row_vals = [row_label] + [f"{grid[r][c]:.3f}" for c in range(12)]
        lines.append(delimiter.join(row_vals))

    encoding = file_format.get('encoding', 'UTF-8')
    try:
        with open(output_path, 'w', encoding=encoding) as f:
            f.write('\n'.join(lines) + '\n')
        print(f"✓ Generated plate grid (8x12) to {output_path}")
        print(f"  Template: {template.get('analyzer', {}).get('name', 'Unknown')}")
        return True
    except IOError as e:
        print(f"Error writing file: {e}", file=sys.stderr)
        return False


def generate_xml(template, output_path, count=5):
    """Generate DT-Prime-style XML from template (package/plate/cell/test/result schema)."""
    protocol = template.get('protocol', {})
    if protocol.get('type') != 'FILE':
        print(f"Error: Template protocol type is '{protocol.get('type')}', expected 'FILE'",
              file=sys.stderr)
        return False

    xml_fmt = template.get('xmlFormat', {})
    encoding = xml_fmt.get('encoding', 'windows-1251')
    protocol_path = xml_fmt.get('structure', {}).get('protocol_path', 'C:\\DT-Prime\\run.r96')
    test_samples = template.get('testSamples', [])

    samples = test_samples[:count] if test_samples else []
    if not samples:
        for i in range(count):
            samples.append({
                'name': f"tst{datetime.now().strftime('%Y%m%d')}{i+1:02d}",
                'x': str((i % 12) + 1),
                'y': str((i // 12) + 1),
                'result': '+' if i % 2 == 0 else '-'
            })

    lines = [
        '<?xml version="1.0" encoding="' + encoding + '" standalone="yes"?>',
        '<package>',
        '  <RealTime_PCR ProtocolPath="' + protocol_path + '"/>',
        '  <data>',
        '    <plate id="0">'
    ]

    for s in samples:
        name = s.get('name', '')
        x = s.get('x', '1')
        y = s.get('y', '1')
        result = s.get('result', '-')
        lines.append(f'      <cell x="{x}" y="{y}" name="{name}" state="complete">')
        lines.append(f'        <test id="ASSAY" value="{result}">')
        lines.append(f'          <result name="qualitative" value="{result}"/>')
        lines.append('        </test>')
        lines.append('      </cell>')

    lines.extend([
        '    </plate>',
        '  </data>',
        '</package>'
    ])

    try:
        with open(output_path, 'w', encoding=encoding) as f:
            f.write('\n'.join(lines) + '\n')
        print(f"✓ Generated {len(samples)} cells to {output_path}")
        print(f"  Template: {template.get('analyzer', {}).get('name', 'Unknown')}")
        print(f"  Format: XML ({encoding})")
        return True
    except (IOError, LookupError) as e:
        print(f"Error writing file: {e}", file=sys.stderr)
        return False


def generate_xlsx(template, output_path, count=5):
    """Generate XLSX file from Excel template (e.g. FluoroCycler XT 12-column)."""
    try:
        from openpyxl import Workbook
    except ImportError:
        print("Error: openpyxl required for XLSX. Install with: pip install openpyxl",
              file=sys.stderr)
        return False

    protocol = template.get('protocol', {})
    if protocol.get('type') != 'FILE':
        print(f"Error: Template protocol type is '{protocol.get('type')}', expected 'FILE'",
              file=sys.stderr)
        return False

    excel_format = template.get('excelFormat', {})
    columns = sorted(template.get('columns', []), key=lambda c: c.get('index', 0))
    test_samples = template.get('testSamples', [])

    if not columns:
        print("Error: Template has no columns defined", file=sys.stderr)
        return False

    wb = Workbook()
    ws = wb.active
    sheet_name = excel_format.get('sheetName', 'Results')
    ws.title = sheet_name

    row_idx = 1

    # Header row
    for col_idx, col in enumerate(columns, start=1):
        ws.cell(row=row_idx, column=col_idx, value=col['name'])
    row_idx += 1

    samples = test_samples[:count] if test_samples else []
    if not samples:
        for i in range(count):
            pos = f"{chr(65 + (i // 12))}{(i % 12) + 1}"
            samples.append({
                'SampleID': f"FC-{datetime.now().strftime('%Y%m%d')}-{i+1:04d}",
                'WellPosition': pos,
                'AssayName': 'FluoroType MTBDR 2.0',
                'TargetName': 'MTB',
                'TargetNo': '0',
                'CP': 14.5 if i % 2 == 0 else -1.0,
                'Interpretation': 'DETECTED' if i % 2 == 0 else 'NOT DETECTED',
                'CalcConc': '',
                'CalcConcUnit': '',
                'RunDate': datetime.now().strftime('%Y-%m-%d'),
                'RunID': '',
                'Notes': ''
            })

    for sample in samples:
        for col_idx, col in enumerate(columns, start=1):
            val = sample.get(col['name'], '')
            ws.cell(row=row_idx, column=col_idx, value=val)
        row_idx += 1

    try:
        wb.save(output_path)
        print(f"✓ Generated {len(samples)} samples to {output_path}")
        print(f"  Template: {template.get('analyzer', {}).get('name', 'Unknown')}")
        print(f"  Format: XLSX, sheet: {sheet_name}")
        return True
    except IOError as e:
        print(f"Error writing file: {e}", file=sys.stderr)
        return False


def _well_position_to_index(pos):
    """Convert A01, B12 etc to 1-96 well index."""
    if not pos or len(pos) < 2:
        return 1
    row = ord(pos[0].upper()) - ord('A')
    try:
        col = int(pos[1:]) - 1
    except ValueError:
        col = 0
    return row * 12 + col + 1


def _write_quantstudio_row(ws, row_idx, columns, well_num, pos, omit, sample_id,
                           target_name, task, reporter, quencher, ct, qty_mean):
    """Write one QuantStudio data row (VIH-1 or IC)."""
    col_map = {c['name']: c['index'] for c in columns}
    defaults = {
        'Well': well_num,
        'Well Position': pos,
        'Omit': omit,
        'Sample Name': sample_id,
        'Target Name': target_name,
        'Task': task,
        'Reporter': reporter,
        'Quencher': quencher,
        'CT': ct,
        'Ct Mean': ct if ct != 'Undetermined' and ct != '' else '',
        'Ct SD': '',
        'Quantity': qty_mean,
        'Quantity Mean': qty_mean,
        'Quantity SD': '',
        'Automatic Ct Threshold': '1',
        'Ct Threshold': '0.2',
        'Automatic Baseline': '1',
        'Baseline Start': '1.0',
        'Baseline End': '50.0',
        'Comments': '',
        'NOAMP': '',
        'EXPFAIL': '',
        'THOLDFAIL': '',
        'HIGHSD': '',
        'PRFLOW': '',
        'CQCONF': '',
        'Amp Score': '',
        'Y-Intercept': '',
        'R²': '',
        'Slope': '',
        'Efficiency': ''
    }
    for col in columns:
        idx = col['index']
        val = defaults.get(col['name'], '')
        if isinstance(val, (int, float)) and val != '':
            ws.write(row_idx, idx, val)
        else:
            ws.write(row_idx, idx, str(val) if val else '')


def main():
    """CLI entry point for file generator."""
    parser = argparse.ArgumentParser(
        description='Generate CSV/TXT files from FILE protocol analyzer templates'
    )
    parser.add_argument(
        '--template', '-t',
        type=str,
        required=True,
        help='Template name (without .json extension)'
    )
    parser.add_argument(
        '--output', '-o',
        type=str,
        required=True,
        help='Output file path'
    )
    parser.add_argument(
        '--count', '-c',
        type=int,
        default=5,
        help='Number of data rows to generate (default: 5)'
    )
    parser.add_argument(
        '--templates-dir',
        type=str,
        help='Templates directory path (optional)'
    )
    parser.add_argument(
        '--format', '-f',
        type=str,
        choices=['csv', 'xls', 'xlsx', 'xml'],
        help='Output format: csv, xls, xlsx, or xml (default: auto-detect from template)'
    )

    args = parser.parse_args()

    # Load template
    try:
        loader = TemplateLoader(args.templates_dir)
        template = loader.load_template(args.template, validate=True)
    except FileNotFoundError:
        print(f"Error: Template '{args.template}' not found", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"Error loading template: {e}", file=sys.stderr)
        return 1

    # Determine format: explicit --format, or from template
    out_format = args.format
    if not out_format:
        file_config = template.get('file_config', {})
        excel_fmt = template.get('excelFormat', {})
        xml_fmt = template.get('xmlFormat', {})
        if file_config.get('format') == 'XML' or xml_fmt:
            out_format = 'xml'
        elif file_config.get('format') == 'XLSX' or excel_fmt.get('format') == 'xlsx':
            out_format = 'xlsx'
        elif file_config.get('format') == 'XLS' or excel_fmt:
            out_format = 'xls'
        else:
            out_format = 'csv'

    if out_format == 'xml':
        success = generate_xml(template, args.output, args.count)
    elif out_format == 'xlsx':
        success = generate_xlsx(template, args.output, args.count)
    elif out_format == 'xls':
        success = generate_xls(template, args.output, args.count)
    else:
        success = generate_csv(template, args.output, args.count)

    return 0 if success else 1


if __name__ == '__main__':
    sys.exit(main())
