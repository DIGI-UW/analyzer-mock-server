"""
Consumer-contract guard for the canonical analyzer profile baseline.

Two levels, both must hold and block in CI (seam, not stub):

  1. SCHEMA CONFORMANCE — every canonical profile validates against
     projects/analyzer-profiles/schema/analyzer-defaults-1.0.schema.json
     (the contract: required profileMeta{id,version,displayName,confidence},
     category, protocol, configDefaults; socket profiles need analyzer_name/
     manufacturer/transport + non-empty default_test_mappings; FILE profiles
     need supported_extensions/column_mapping; every mapping has non-empty
     test_code+loinc — the exact keys OE2 AnalyzerService parses).

  2. REAL-ADAPTER LOAD — every profile-backed mock template loads through the
     actual profile_adapter.load_profile_backed_template (the real reader the
     mock uses) and yields assay fields that each carry a LOINC. This proves the
     normalized profiles are consumable by the real code path, not just that a
     JSON validator likes them.

This is the RED->GREEN gate for the profile-format normalization workstream.
"""

import glob
import json
import os

import pytest
from jsonschema import Draft7Validator

import profile_adapter

PROFILES_ROOT = profile_adapter._profiles_root()
SCHEMA_PATH = os.path.join(PROFILES_ROOT, "schema", "analyzer-defaults-1.0.schema.json")
TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")


def _load(path):
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def _profile_files():
    """All canonical profile JSONs (astm/ hl7/ file/), excluding the schema dir."""
    out = []
    for path in sorted(glob.glob(os.path.join(PROFILES_ROOT, "*", "*.json"))):
        if os.sep + "schema" + os.sep in path:
            continue
        out.append(path)
    return out


def test_baseline_corpus_present():
    """Fail LOUDLY (never silently skip) if the profile corpus/schema isn't on
    disk — this guard is meaningful only in the integrated checkout (OE2 with the
    mock submodule) or when ANALYZER_PROFILES_DIR points at the profiles."""
    assert os.path.isfile(SCHEMA_PATH), (
        f"canonical schema not found at {SCHEMA_PATH}; run from the integrated "
        f"checkout or set ANALYZER_PROFILES_DIR"
    )
    assert _profile_files(), f"no canonical profiles found under {PROFILES_ROOT}"


def _profile_backed_templates():
    names = []
    for path in sorted(glob.glob(os.path.join(TEMPLATES_DIR, "*.json"))):
        if os.path.basename(path) == "schema.json":
            continue
        if _load(path).get("profile"):
            names.append(os.path.basename(path))
    return names


@pytest.fixture(scope="module")
def validator():
    return Draft7Validator(_load(SCHEMA_PATH))


@pytest.mark.parametrize(
    "profile_path", _profile_files(),
    ids=[os.path.relpath(p, PROFILES_ROOT) for p in _profile_files()],
)
def test_profile_conforms_to_canonical_schema(profile_path, validator):
    profile = _load(profile_path)
    errors = sorted(validator.iter_errors(profile), key=lambda e: list(e.path))
    detail = "\n".join(f"    {list(e.path)}: {e.message}" for e in errors)
    assert not errors, (
        f"{os.path.relpath(profile_path, PROFILES_ROOT)} violates canonical schema:\n{detail}"
    )


TEMPLATE_SCHEMA_PATH = os.path.join(TEMPLATES_DIR, "schema.json")


def _template_files():
    return [
        p for p in sorted(glob.glob(os.path.join(TEMPLATES_DIR, "*.json")))
        if os.path.basename(p) != "schema.json"
    ]


@pytest.fixture(scope="module")
def template_validator():
    return Draft7Validator(_load(TEMPLATE_SCHEMA_PATH))


@pytest.mark.parametrize(
    "template_path", _template_files(),
    ids=[os.path.basename(p) for p in _template_files()],
)
def test_template_conforms_to_schema(template_path, template_validator):
    tpl = _load(template_path)
    errors = sorted(template_validator.iter_errors(tpl), key=lambda e: list(e.path))
    detail = "\n".join(f"    {list(e.path)}: {e.message}" for e in errors)
    assert not errors, (
        f"{os.path.basename(template_path)} violates template schema:\n{detail}"
    )


@pytest.mark.parametrize("template_name", _profile_backed_templates())
def test_profile_backed_template_loads_through_real_adapter(template_name):
    tpl = _load(os.path.join(TEMPLATES_DIR, template_name))
    merged = profile_adapter.load_profile_backed_template(template_name, tpl)
    assert merged is not None, f"{template_name}: adapter returned None (no profile?)"
    fields = merged.get("fields", [])
    assert fields, f"{template_name}: profile yielded zero assay fields"
    missing_loinc = [f.get("code") for f in fields if not f.get("loinc")]
    assert not missing_loinc, f"{template_name}: assay fields missing LOINC: {missing_loinc}"


# --- Fixture faithfulness: parse each loadable fixture through a faithful mirror
# of the bridge FileResultParser contract and assert it would import (≥1 result
# whose testCode maps to a LOINC), and that it carries zero real-PHI patterns. ---

import csv as _csv
import re as _re

_OE2_ROOT = os.path.dirname(os.path.dirname(PROFILES_ROOT))
FIXTURES_DIR = os.path.join(_OE2_ROOT, "src", "test", "resources", "testdata", "files")

# (fixture filename, profile relpath under projects/analyzer-profiles)
FILE_FIXTURES = [
    ("tecan-f50-results.csv", "file/tecan-f50.json"),
    ("multiskan-fc-results.csv", "file/multiskan-fc.json"),
    ("fluorocycler-results.xlsx", "file/fluorocycler-xt.json"),
    ("quantstudio5-results.xlsx", "file/quantstudio.json"),
    ("quantstudio7-results.xlsx", "file/quantstudio.json"),
]

# Real Madagascar accession/identifier shapes that must never appear in a fixture.
_PHI_PATTERNS = [_re.compile(p) for p in (r"\bLM\d{6,}", r"CG-M4-\d", r"\bLL\d{6,}", r"\bDCN\d{6,}")]


def _rows_from_fixture(path):
    """Return list-of-rows (each a list of cell strings) for csv/xlsx, mirroring
    the bridge's reader dispatch."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv", ".txt"):
        with open(path, "r", encoding="utf-8-sig", newline="") as fh:
            sample = fh.readline()
            delim = ";" if sample.count(";") >= sample.count(",") else ","
            fh.seek(0)
            return [[(c or "").strip() for c in row] for row in _csv.reader(fh, delimiter=delim)]
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Results"] if "Results" in wb.sheetnames else wb.worksheets[0]
    return [[("" if c is None else str(c)).strip() for c in row] for row in ws.iter_rows(values_only=True)]


def _find_header_row(rows):
    """Mirror FileResultParser.findHeaderRow (sheet/CSV header location)."""
    if not rows:
        return -1
    a0 = rows[0][0] if rows[0] else ""
    if a0 == "Well":
        return 0
    if "Block Type" in a0 or "Experiment Name" in a0:
        for i in range(min(200, len(rows))):
            if rows[i] and rows[i][0] == "Well":
                return i
        return 0
    return 0


def _emit(rows, column_mapping):
    """Mirror the bridge: map header→role, emit a row iff sampleId+testCode+value
    (result→ctValue→interpretation) are all non-blank."""
    h = _find_header_row(rows)
    header = rows[h]
    role_idx = {}
    for i, name in enumerate(header):
        role = column_mapping.get(name)
        if role:
            role_idx[role] = i

    def cell(row, role):
        i = role_idx.get(role)
        return row[i] if (i is not None and i < len(row)) else ""

    out = []
    for row in rows[h + 1:]:
        if not any(row):
            continue
        sample_id = cell(row, "sampleId")
        test_code = cell(row, "testCode")
        value = cell(row, "result") or cell(row, "ctValue") or cell(row, "interpretation")
        if sample_id and test_code and value:
            out.append((sample_id, test_code, value))
    return out


@pytest.mark.parametrize(
    "fixture_name,profile_rel", FILE_FIXTURES,
    ids=[f[0] for f in FILE_FIXTURES],
)
def test_fixture_parses_and_is_phi_free(fixture_name, profile_rel):
    fixture = os.path.join(FIXTURES_DIR, fixture_name)
    assert os.path.isfile(fixture), f"loadable fixture missing: {fixture}"
    profile = _load(os.path.join(PROFILES_ROOT, profile_rel))
    column_mapping = profile.get("column_mapping", {})
    mapped_codes = {m.get("test_code") for m in profile.get("default_test_mappings", [])}

    emitted = _emit(_rows_from_fixture(fixture), column_mapping)
    assert emitted, f"{fixture_name}: parses to zero importable rows (sampleId+testCode+value)"
    patient = [(s, t, v) for (s, t, v) in emitted if t in mapped_codes]
    assert patient, (
        f"{fixture_name}: no row whose testCode is in the profile's default_test_mappings "
        f"{sorted(mapped_codes)}; emitted testCodes={sorted({t for _, t, _ in emitted})}"
    )

    blob = "\n".join("\t".join(r) for r in _rows_from_fixture(fixture))
    leaks = [p.pattern for p in _PHI_PATTERNS if p.search(blob)]
    assert not leaks, f"{fixture_name}: real-PHI identifier pattern(s) present: {leaks}"


def test_validated_file_profiles_have_faithful_fixture():
    """VALIDATED gate: a FILE profile may not claim confidence=VALIDATED without a
    faithful loadable fixture behind it (no 'validated' label on untested data)."""
    fixtured = {pr for _, pr in FILE_FIXTURES}
    offenders = []
    for path in _profile_files():
        prof = _load(path)
        if prof.get("protocol", {}).get("name") != "FILE":
            continue
        if prof.get("profileMeta", {}).get("confidence") != "VALIDATED":
            continue
        rel = os.path.relpath(path, PROFILES_ROOT).replace(os.sep, "/")
        if rel not in fixtured:
            offenders.append(rel)
    assert not offenders, (
        f"VALIDATED FILE profiles with no faithful fixture (downgrade confidence or add a "
        f"fixture): {offenders}"
    )
