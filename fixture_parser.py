"""
Fixture file parser — extracts metadata (accessions, results) from real analyzer export files.

Supports CSV, XLSX (.xlsx via openpyxl), and XLS (.xls via xlrd).
Used by the mock server to return parsed metadata alongside dropped fixture files,
so E2E tests never need to hardcode expected values.

Control row filtering uses the same prefixes as the bridge's FileResultParser.
"""

import csv
import io
import logging
import os
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

CONTROL_PREFIXES = [
    # Molecular (QuantStudio, FluoroCycler)
    "CNEG", "CPOS", "NTC", "PTC",
    # ELISA plate readers (Tecan, Multiskan)
    "NEG", "POS", "NC", "PC", "BLANC", "BLANK",
]


def _is_control(sample_id: str) -> bool:
    upper = sample_id.upper().strip()
    return any(upper.startswith(prefix) for prefix in CONTROL_PREFIXES)


def parse_fixture(fixture_path: str, fixture_config: Dict[str, Any]) -> List[Dict[str, str]]:
    """Parse a fixture file and return patient (non-control) results.

    Args:
        fixture_path: Absolute path to the fixture file.
        fixture_config: Dict with keys:
            - format: "CSV" | "XLSX" | "XLS"
            - column_mapping: { "sampleId": "<header>", "result": "<header>", "testCode": "<header>" }
            - delimiter: str (CSV only, default ",")
            - skipRows: int (CSV only, default 0)

    Returns:
        List of dicts: [{"sampleId": "...", "result": "...", "testCode": "..."}, ...]
    """
    fmt = (fixture_config.get("format") or "CSV").upper()
    col_map = fixture_config.get("column_mapping", {})
    test_code_filter = fixture_config.get("testCodeFilter")

    if fmt == "CSV":
        results = _parse_csv(fixture_path, col_map, fixture_config)
    elif fmt == "XLSX":
        results = _parse_xlsx(fixture_path, col_map)
    elif fmt == "XLS":
        results = _parse_xls(fixture_path, col_map)
    else:
        raise ValueError(f"Unsupported fixture format: {fmt}")

    if test_code_filter:
        results = [r for r in results if r.get("testCode") == test_code_filter]

    return results


def _parse_csv(
    path: str,
    col_map: Dict[str, str],
    config: Dict[str, Any],
) -> List[Dict[str, str]]:
    delimiter = config.get("delimiter", ",")
    skip_rows = int(config.get("skipRows", 0))

    sample_col = col_map.get("sampleId", "Sample Name")
    result_col = col_map.get("result", "Result")
    test_col = col_map.get("testCode")

    with open(path, "r", encoding="utf-8-sig") as f:
        lines = f.readlines()

    if len(lines) <= skip_rows:
        raise ValueError(f"Fixture {os.path.basename(path)} has {len(lines)} lines but skipRows={skip_rows}")

    # Skip metadata rows, then parse with csv module
    data_lines = lines[skip_rows:]
    reader = csv.DictReader(io.StringIO("".join(data_lines)), delimiter=delimiter)

    results = []
    for row in reader:
        sample_id = (row.get(sample_col) or "").strip()
        result_val = (row.get(result_col) or "").strip()

        if not sample_id or not result_val:
            continue
        if _is_control(sample_id):
            continue

        entry = {"sampleId": sample_id, "result": result_val}
        if test_col and row.get(test_col):
            entry["testCode"] = row[test_col].strip()
        results.append(entry)

    return results


def _parse_xlsx(path: str, col_map: Dict[str, str]) -> List[Dict[str, str]]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for .xlsx parsing: pip install openpyxl")

    sample_col = col_map.get("sampleId", "Sample Name")
    result_col = col_map.get("result", "Result")
    test_col = col_map.get("testCode")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Header-scan: find first sheet + row containing the sampleId column header
    header_row_idx = None
    ws = None
    col_indices = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row_idx, row in enumerate(ws.iter_rows(max_row=50, values_only=True), start=1):
            if row and sample_col in [str(c).strip() if c else "" for c in row]:
                header_row_idx = row_idx
                headers = [str(c).strip() if c else "" for c in row]
                for i, h in enumerate(headers):
                    if h == sample_col:
                        col_indices["sampleId"] = i
                    if h == result_col:
                        col_indices["result"] = i
                    if test_col and h == test_col:
                        col_indices["testCode"] = i
                break
        if header_row_idx:
            break

    if not header_row_idx or "sampleId" not in col_indices or "result" not in col_indices:
        wb.close()
        raise ValueError(f"Headers '{sample_col}' / '{result_col}' not found in {os.path.basename(path)}")

    results = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx <= header_row_idx:
            continue
        if not row:
            continue

        sample_id = str(row[col_indices["sampleId"]] or "").strip()
        result_val = str(row[col_indices["result"]] or "").strip()

        if not sample_id or not result_val:
            continue
        if _is_control(sample_id):
            continue

        entry = {"sampleId": sample_id, "result": result_val}
        if "testCode" in col_indices and row[col_indices["testCode"]]:
            entry["testCode"] = str(row[col_indices["testCode"]]).strip()
        results.append(entry)

    wb.close()
    return results


def _parse_xls(path: str, col_map: Dict[str, str]) -> List[Dict[str, str]]:
    try:
        import xlrd
    except ImportError:
        raise ImportError("xlrd is required for .xls parsing: pip install xlrd")

    sample_col = col_map.get("sampleId", "Sample Name")
    result_col = col_map.get("result", "Result")
    test_col = col_map.get("testCode")

    wb = xlrd.open_workbook(path)

    # Header-scan across sheets
    header_row_idx = None
    ws = None
    col_indices = {}

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        for row_idx in range(min(ws.nrows, 50)):
            row = [str(ws.cell_value(row_idx, c)).strip() for c in range(ws.ncols)]
            if sample_col in row:
                header_row_idx = row_idx
                for i, h in enumerate(row):
                    if h == sample_col:
                        col_indices["sampleId"] = i
                    if h == result_col:
                        col_indices["result"] = i
                    if test_col and h == test_col:
                        col_indices["testCode"] = i
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None or "sampleId" not in col_indices or "result" not in col_indices:
        raise ValueError(f"Headers '{sample_col}' / '{result_col}' not found in {os.path.basename(path)}")

    results = []
    for row_idx in range(header_row_idx + 1, ws.nrows):
        sample_id = str(ws.cell_value(row_idx, col_indices["sampleId"])).strip()
        result_val = str(ws.cell_value(row_idx, col_indices["result"])).strip()

        if not sample_id or not result_val:
            continue
        if _is_control(sample_id):
            continue

        entry = {"sampleId": sample_id, "result": result_val}
        if "testCode" in col_indices:
            val = str(ws.cell_value(row_idx, col_indices["testCode"])).strip()
            if val:
                entry["testCode"] = val
        results.append(entry)

    return results
